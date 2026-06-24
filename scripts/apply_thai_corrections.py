# -*- coding: utf-8 -*-
"""Apply Thai copy corrections from Website Project.xlsx (column F) to the site."""
from __future__ import annotations

import json
import re
import unicodedata
from html import escape as html_escape
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = Path(r"c:\Users\MSI\Downloads\Website Project.xlsx")
I18N = ROOT / "js" / "i18n-data.js"
INDEX = ROOT / "index.html"

# Keys whose values may contain HTML markup.
HTML_KEYS = {
    "hero_title",
    "bbg_title",
    "about_title",
    "tech_title",
    "fac_title",
    "process_title",
    "contact_title",
    "contact_hours",
    "footer_addr",
    "footer_clinic_name",
    "footer_rep",
    "faq5_a",
}

THAI_RE = re.compile(r"[\u0e00-\u0e7f]")
HANGUL_RE = re.compile(r"[\uac00-\ud7af]")

EMOJI_RE = re.compile(
    "["
    "\U0001F300-\U0001FAFF"
    "\U00002600-\U000027BF"
    "\U0000FE0F"
    "\U0000200D"
    "]+",
    flags=re.UNICODE,
)


def should_apply(value: str, key: str) -> bool:
    if HANGUL_RE.search(value):
        return False
    if THAI_RE.search(value):
        return True
    # Allow short language-neutral UI tokens (numbers, labels).
    return bool(re.fullmatch(r"[\d:+\-./A-Za-z ]+", value))


def clean_text(value: str, key: str) -> str:
    text = value.replace("\r\n", "\n").strip()
    text = EMOJI_RE.sub("", text)
    text = re.sub(r"https?://\S+", "", text)
    text = re.sub(r"\[image[^\]]*\]", "", text, flags=re.I)
    text = re.sub(r" +", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text).strip()

    if key == "meta_description":
        text = " ".join(line.strip() for line in text.splitlines() if line.strip())
    elif key in HTML_KEYS:
        text = text.replace("\n", "<br>")

    text = re.sub(r"EVERM\s+DENTAL\s+CLINIC", "EVERM SURGERY CLINIC", text, flags=re.I)
    text = re.sub(r"EVERM\s+Dental\s+Clinic", "EVERM Surgery Clinic", text, flags=re.I)
    text = re.sub(r"DENTAL\s+CLINIC", "SURGERY CLINIC", text, flags=re.I)
    text = re.sub(r"Dental\s+Clinic", "Surgery Clinic", text, flags=re.I)

    return text


def load_corrections() -> dict[str, str]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Thai (th)"]
    out: dict[str, str] = {}
    for row in range(2, ws.max_row + 1):
        key = ws.cell(row, 3).value
        new = ws.cell(row, 6).value
        if not key or new is None:
            continue
        key = str(key).strip()
        text = clean_text(str(new), key)
        if not text or not should_apply(text, key):
            continue
        out[key] = text
    return out


def js_quote(value: str) -> str:
    return json.dumps(value, ensure_ascii=False)


def update_i18n(corrections: dict[str, str]) -> list[str]:
    text = I18N.read_text(encoding="utf-8")
    start = text.index("  th: {")
    end = text.index("\n  }\n};", start)
    th_block = text[start:end]

    updated: list[str] = []
    for key, value in corrections.items():
        pattern = rf"(\n    {re.escape(key)}: )(\"(?:\\.|[^\"\\])*\")"
        match = re.search(pattern, th_block)
        if not match:
            updated.append(f"MISSING:{key}")
            continue
        old_quoted = match.group(2)
        old_value = json.loads(old_quoted)
        if old_value == value:
            continue
        th_block = th_block[: match.start(2)] + js_quote(value) + th_block[match.end(2) :]
        updated.append(key)

    new_text = text[:start] + th_block + text[end:]
    I18N.write_text(new_text, encoding="utf-8")
    return updated


def _sub_inner(pattern: str, html: str, value: str) -> tuple[str, int]:
    def repl(match: re.Match[str]) -> str:
        return match.group(1) + value + match.group(3)

    return re.subn(pattern, repl, html, count=1, flags=re.S)


def update_index_html(corrections: dict[str, str]) -> int:
    html = INDEX.read_text(encoding="utf-8")
    count = 0

    for key, value in corrections.items():
        if key in HTML_KEYS or ("<" in value and ">" in value):
            pattern = rf'(<[^>]+data-i18n-html="{re.escape(key)}"[^>]*>)(.*?)(</)'
            html, n = _sub_inner(pattern, html, value)
            if n:
                count += n
                continue
            pattern = rf'(<[^>]+data-i18n="{re.escape(key)}"[^>]*>)(.*?)(</)'
            if "<" in value:
                html, n = _sub_inner(pattern, html, value)
                if n:
                    count += n
                    continue

        for attr, attr_name in (
            ("data-i18n-placeholder", "placeholder"),
            ("data-i18n-aria", "aria-label"),
            ("data-i18n-alt", "alt"),
            ("data-i18n-title", "title"),
        ):
            pattern = rf'(<[^>]*{attr}="{re.escape(key)}"[^>]*)(>)'
            match = re.search(pattern, html)
            if not match:
                continue
            tag = match.group(1)
            tag = re.sub(rf'\s{attr_name}="[^"]*"', "", tag)
            tag = f'{tag} {attr_name}="{html_escape(value, quote=True)}"'
            html = html[: match.start()] + tag + match.group(2) + html[match.end() :]
            count += 1
            break
        else:
            pattern = rf'(<[^>]+data-i18n="{re.escape(key)}"[^>]*>)(.*?)(</)'
            match = re.search(pattern, html, flags=re.S)
            if match:
                html, n = _sub_inner(pattern, html, value)
                if n:
                    count += n

        if key == "meta_description":
            html, n = re.subn(
                r'(<meta name="description" content=")([^"]*)(")',
                lambda m: m.group(1) + value + m.group(3),
                html,
                count=1,
            )
            if n:
                count += n
        if key == "meta_title":
            html, n = re.subn(
                r"(<title>)(.*?)(</title>)",
                lambda m: m.group(1) + value + m.group(3),
                html,
                count=1,
                flags=re.S,
            )
            if n:
                count += n

    INDEX.write_text(html, encoding="utf-8")
    return count


def main() -> None:
    if not XLSX.is_file():
        raise SystemExit(f"Excel not found: {XLSX}")

    corrections = load_corrections()
    updated_keys = update_i18n(corrections)

    from sync_index_thai_fallbacks import sync_index, load_th_pack

    html_count = sync_index(load_th_pack())

    report = {
        "corrections_loaded": len(corrections),
        "i18n_updated": len([k for k in updated_keys if not k.startswith("MISSING")]),
        "i18n_missing": [k for k in updated_keys if k.startswith("MISSING")],
        "html_patches": html_count,
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
