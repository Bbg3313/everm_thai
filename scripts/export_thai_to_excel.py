#!/usr/bin/env python3
"""Export all Thai (th) strings from js/i18n-data.js to Excel."""

import re
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
I18N_FILE = ROOT / "js" / "i18n-data.js"
OUT_FILE = ROOT / "everm-thai-strings.xlsx"


def section_for_key(key: str) -> str:
    if key.startswith("meta_"):
        return "SEO / Meta"
    if key.startswith(("nav_", "skip_", "logo_", "lang_", "contact_marquee")):
        return "Navigation"
    if key.startswith("hero_"):
        return "Hero"
    if key.startswith(("trust_", "bbg_")):
        return "Trust / BlueBridge"
    if key.startswith("about_") or key.startswith("feat"):
        return "About"
    if key.startswith(("director_video", "patient_video")):
        return "Videos"
    if key.startswith(("services_", "svc", "treatments_")):
        return "Services"
    if key.startswith("tech_"):
        return "Technology"
    if key.startswith("fac_"):
        return "Facilities"
    if key.startswith(("process_", "step")):
        return "Process"
    if key.startswith(("team_", "doctor_", "director_", "specialist_")):
        return "Team"
    if key.startswith("case"):
        return "Facility highlights"
    if key.startswith("review"):
        return "Reviews"
    if key.startswith("faq_"):
        return "FAQ"
    if key.startswith(("contact_", "form_")):
        return "Contact / Form"
    if key.startswith("footer_"):
        return "Footer"
    if key.startswith("float_"):
        return "Floating CTA"
    return "Other"


def strip_html(text: str) -> str:
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    text = text.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">").replace("&quot;", '"')
    return text.strip()


def parse_th_block(content: str) -> list[tuple[str, str]]:
    start = content.find("  th: {")
    if start < 0:
        raise SystemExit("th: block not found in i18n-data.js")
    depth = 0
    i = content.index("{", start)
    begin = i
    for j in range(i, len(content)):
        if content[j] == "{":
            depth += 1
        elif content[j] == "}":
            depth -= 1
            if depth == 0:
                block = content[begin + 1 : j]
                break
    else:
        raise SystemExit("Could not parse th block")

    pattern = re.compile(
        r'^\s*([a-zA-Z0-9_]+):\s*(?:"((?:\\.|[^"\\])*)"|\'((?:\\.|[^\'\\])*)\')\s*,?\s*$',
        re.M,
    )
    rows = []
    for m in pattern.finditer(block):
        key = m.group(1)
        raw = m.group(2) if m.group(2) is not None else m.group(3)
        raw = raw.replace("\\n", "\n").replace('\\"', '"').replace("\\'", "'")
        rows.append((key, raw))
    return rows


def main():
    content = I18N_FILE.read_text(encoding="utf-8")
    entries = parse_th_block(content)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Thai (th)"

    headers = ["No.", "Section", "Key", "Thai (plain text)", "Thai (with HTML)"]
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    for idx, (key, raw) in enumerate(entries, 1):
        ws.append([idx, section_for_key(key), key, strip_html(raw), raw])

    ws.freeze_panes = "A2"
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        if col == 1:
            ws.column_dimensions[letter].width = 6
        elif col == 2:
            ws.column_dimensions[letter].width = 18
        elif col == 3:
            ws.column_dimensions[letter].width = 28
        else:
            ws.column_dimensions[letter].width = 55
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    meta = wb.create_sheet("Info")
    meta.append(["Source", str(I18N_FILE.relative_to(ROOT))])
    meta.append(["Locale", "th"])
    meta.append(["Total strings", len(entries)])
    meta.append(["Note", "Plain text column strips <br> and HTML tags for easier editing."])

    wb.save(OUT_FILE)
    print(f"Wrote {len(entries)} rows to {OUT_FILE}")


if __name__ == "__main__":
    main()
