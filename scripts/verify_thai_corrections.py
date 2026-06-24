# -*- coding: utf-8 -*-
"""Compare Website Project.xlsx column F with js/i18n-data.js th pack."""
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = Path(r"c:\Users\MSI\Downloads\Website Project.xlsx")
I18N = ROOT / "js" / "i18n-data.js"

HTML_KEYS = {
    "hero_title",
    "bbg_title",
    "about_title",
    "tech_title",
    "fac_title",
    "process_title",
    "contact_title",
    "contact_hours",
    "footer_addr",
    "footer_clinic_name",
    "footer_rep",
    "faq5_a",
}
THAI_RE = re.compile(r"[\u0e00-\u0e7f]")
HANGUL_RE = re.compile(r"[\uac00-\ud7af]")
EMOJI_RE = re.compile(
    "["
    "\U0001F300-\U0001FAFF"
    "\U00002600-\U000027BF"
    "\U0000FE0F"
    "\U0000200D"
    "]+",
    flags=re.UNICODE,
)


def clean_text(value: str, key: str) -> str:
    text = str(value).replace("\r\n", "\n").strip()
    text = EMOJI_RE.sub("", text)
    text = re.sub(r"https?://\S+", "", text)
    text = re.sub(r"\[image[^\]]*\]", "", text, flags=re.I)
    text = re.sub(r" +", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    if key == "meta_description":
        text = " ".join(line.strip() for line in text.splitlines() if line.strip())
    elif key in HTML_KEYS:
        text = text.replace("\n", "<br>")
    text = re.sub(r"EVERM\s+DENTAL\s+CLINIC", "EVERM SURGERY CLINIC", text, flags=re.I)
    text = re.sub(r"EVERM\s+Dental\s+Clinic", "EVERM Surgery Clinic", text, flags=re.I)
    text = re.sub(r"DENTAL\s+CLINIC", "SURGERY CLINIC", text, flags=re.I)
    text = re.sub(r"Dental\s+Clinic", "Surgery Clinic", text, flags=re.I)
    return text


def should_apply(value: str) -> bool:
    if HANGUL_RE.search(value):
        return False
    if THAI_RE.search(value):
        return True
    return bool(re.fullmatch(r"[\d:+\-./A-Za-z ]+", value))


def load_excel() -> tuple[dict[str, str], dict[str, dict], list[str]]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Thai (th)"]
    applicable: dict[str, str] = {}
    skipped: dict[str, dict] = {}
    empty_f: list[str] = []
    for row in range(2, ws.max_row + 1):
        key = ws.cell(row, 3).value
        new = ws.cell(row, 6).value
        if not key:
            continue
        key = str(key).strip()
        if new is None or str(new).strip() == "":
            empty_f.append(key)
            continue
        text = clean_text(new, key)
        if not text:
            skipped[key] = {"reason": "empty_after_clean", "raw": str(new)[:200]}
        elif not should_apply(text):
            skipped[key] = {"reason": "not_thai_or_note", "raw": str(new)[:200]}
        else:
            applicable[key] = text
    return applicable, skipped, empty_f


def load_site_th() -> dict[str, str]:
    text = I18N.read_text(encoding="utf-8")
    start = text.index("  th: {")
    end = text.index("\n  }\n};", start)
    block = text[start:end]
    pack: dict[str, str] = {}
    pattern = re.compile(r'\n    ([a-zA-Z0-9_]+): ("(?:\\.|[^"\\])*")')
    for match in pattern.finditer(block):
        pack[match.group(1)] = json.loads(match.group(2))
    return pack


def main() -> None:
    excel, skipped, empty_f = load_excel()
    site = load_site_th()

    matched: list[str] = []
    mismatched: list[dict] = []
    missing_in_site: list[str] = []

    for key, expected in excel.items():
        if key not in site:
            missing_in_site.append(key)
        elif site[key] == expected:
            matched.append(key)
        else:
            mismatched.append({"key": key, "excel": expected, "site": site[key]})

    report = {
        "excel_rows_with_F_value": len(excel) + len(skipped),
        "excel_F_empty": len(empty_f),
        "applicable_after_filters": len(excel),
        "matched_exactly": len(matched),
        "mismatched": mismatched,
        "missing_in_site": missing_in_site,
        "skipped_intentionally": skipped,
        "excel_F_empty_keys_sample": empty_f[:30],
    }
    out = ROOT / "scripts" / "_thai_verify.json"
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({k: report[k] for k in report if k != "mismatched"}, ensure_ascii=False, indent=2))
    print("mismatched_count", len(mismatched))


if __name__ == "__main__":
    main()
